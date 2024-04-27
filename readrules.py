#!/usr/bin/env python3
import csv
from os import error
import openpyxl
import sys
from openpyxl.styles import Alignment

# Get the policy csv filename
try:
    csv_file = sys.argv[1]
except:
    print("Please provide the policy csv file as an argument")
    exit (1)

# Open and read the csv file
rulebase = []
with open(csv_file, mode='r') as csvfile:
    rulereader = csv.reader(csvfile)
    for i,row in enumerate(rulereader, start=1):
        rulebase.append(row)

# Create new excel workbook and sheet
wbwr = openpyxl.Workbook()
ws = wbwr.active
ws.title = "Rulebase"

# Copy the rules items in sheet cells
for i,line in enumerate(rulebase, start=1):
    if i != 1:
        line.pop()
    for j,colitem in enumerate(line, start=1):
        ws.cell(row=i,column=j,value=colitem)

# Create styles for alignement that will be applied per cell
allign_vert = Alignment(vertical="center", wrap_text=True)
title_center = Alignment(horizontal="center", vertical="center")

# Go through the rows and mark the section lines differently
# also make sure the first line cells are centered 
# and that the rest of the lines apply word wrap
for i,ws_row in enumerate(ws, start=1):
    if ws.cell(row=i,column=2).value == "Section":
        paint_section = True
    else:
        paint_section = False
    for cell in ws_row:
        cell.alignment=allign_vert
        if paint_section:
            cell.style="Accent1"
        elif i==1:
            cell.alignment=title_center

# Create auto filters defined by the sheet dimensions
ws.auto_filter.ref = ws.dimensions
# Give columns a custom width according to needs
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 34
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['G'].width = 22
ws.column_dimensions['K'].width = 23

# Replace the extension and store new filename for xlsx
# and save and close file
excel_file = csv_file.replace(".csv", ".xlsx")
wbwr.save(excel_file)
print ("Done")
